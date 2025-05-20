import inspect
import logging
from openpyxl import Workbook, load_workbook


class Utils():
    def assertlistitems(self, list, values):
        for stop in list:
            stop_text = stop.text.strip()
            print("The text is: " + stop_text)

            if stop_text.startswith(values):
                print("assert pass")
            else:
                print("FAILED")
                assert False, f"Unexpected stop value: {stop_text}"

    def custom_loggger(self, logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]

        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        fh = logging.FileHandler('automation.log')

        formatter = logging.Formatter('%(asctime)s - %(levelname)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')

        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename= file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        col_count = sh.max_column

        for i in range(2, row_count +1):
            row = []
            for j in range(1, col_count + 1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        # Create an empty list
        datalist = []
        # Open CSV file
        csvdata = open(filename, "r")
        # Create CSV reader
        reader = csv.reader(csvdata)
        # skip header
        next(reader)
        # Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist

